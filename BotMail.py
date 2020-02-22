import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import email.mime.application
import sys
import openpyxl
import time
import re #regex


unlock = 'test'
smtpObj = ""

contactList = []


class Contact:
    def __init__(self, name, email, company):
        self.name = name
        self.email = email
        self.company = company


# method for reading in lines


def unlock_program():
    pw = input('\033[1m' + "Enter password (the word 'test'): ")
    if pw != unlock:
        sys.exit()
    print(
        'Please be sure to read the instructions prior to using this program. \nIf you run into any issues, '
        'please do not hesitate to contact me at k.wu@utexas.edu\n\n')


def login(user, pword):
    global smtpObj
    #special case @utexas.edu
    if "utexas.com" in user.lower():
        smtp_ssl_host = 'smtp.gmail.com'
        emailDomain = 587
    elif "gmail.com" in user.lower():
        smtp_ssl_host = 'smtp.gmail.com'
        emailDomain = 587
    elif "yahoo.com" in user.lower():
        smtp_ssl_host = 'smtp.mail.yahoo.com'
        emailDomain = 465
    elif "outlook.com" in user.lower():
        smtp_ssl_host = 'smtp-mail.outlook.com'
        emailDomain = 587
    elif "hotmail.com" in user.lower():
        smtp_ssl_host = 'smtp-mail.outlook.com'
        emailDomain = 25
    else:
        print("\nEmail domain not detected. TLS encryption failed.")
        time.sleep(1)
        print("\nTerminating program...")
        sys.exit()

    smtpObj = smtplib.SMTP(smtp_ssl_host, emailDomain)
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login(user, pword)



def send_email(user_name):
    global smtpObj
    temp_user_name = user_name
    sender_name = input("Enter sender's name: ")
    subject_line = input("Enter Subject Line: ")
    header = input("Enter salutation: ")
    paragraphs = eval(input("Enter number of paragraphs in email body: "))
    counter = paragraphs
    text = ''
    while counter > 0:
        txt = input("Enter paragraph of text: ")
        text = text + '<p>' + txt + '</p>'
        counter -= 1
    closure = input("Enter closure statement (eg: Best regards, Sincerely, Looking forward to hearing from you soon): ")
    if ',' not in closure:
        closure += ','
    signature = input("Enter your name/signature: ")
    title = input("Enter title (eg: VP Internal): ")


    send__test_email(temp_user_name, subject_line, header, text, closure, signature,title)  # first confirmation

    ready_check = input("Are you absolutely sure the email is ready to be sent out? Enter 'yes'. Otherwise, "
                        "enter 'no' to exit and edit email")
    while True:
        while ready_check.lower() not in ('yes', 'no'):
            ready_check = input("Invalid response. Please enter 'yes' if everything looks good. Otherwise enter 'no' "
                                "to exit program")
        if ready_check.lower() == 'no':
            print("\nTerminating program..... \nPlease fix the text and restart program.")
            time.sleep(3)
            sys.exit()
        if ready_check.lower() == 'yes':
            break

    sent_count = 0

    for x in contactList:
        msg = MIMEMultipart()

        msg['Subject'] = subject_line
        msg['From'] = sender_name

        msg['To'] = x.email

        temporary_name = x.name
        company = x.company  # if needed

        # jank shit time
        temp_text = re.sub(r'\bCOMPANY\b', company, text)
        #
        html = """\
                <html>
                  <head></head>
                  <body>
                    <p>{header} {temporary_name},<br>
                       {temp_text}
                    </p>
                    <br>{closure}<br>{signature}<br>{title}
                  </body>
                </html>
                """.format(**locals())

        msg.attach(MIMEText(html, 'html'))

        # attachment
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open("Coolist_Flyer.pdf", "rb").read())

        encoders.encode_base64(part)

        part.add_header('Content-Disposition', 'attachment; filename="Coolist_Flyer.pdf"')
        msg.attach(part)

        smtpObj.sendmail(user_name, x.email, msg.as_string())
        sent_count += 1

    return sent_count


def send__test_email(temp_user_name, subject_line, header, text, closure, signature, title):
    global smtpObj
    msg = MIMEMultipart()

    msg['Subject'] = subject_line
    msg['From'] = temp_user_name
    msg['To'] = temp_user_name

    temporary_name = temp_user_name
    company = "COMPANYNAMEWILLBEHERE"

    # jank shit time
    temp_text = re.sub(r'\bCOMPANY\b', company, text)
    #

    html = """\
            <html>
              <head></head>
              <body>
                <p>{header} {temporary_name},<br>
                   {temp_text}
                </p>
                <br>{closure}<br>{signature}<br>{title}
              </body>
            </html>
            """.format(**locals())

    msg.attach(MIMEText(html, 'html'))

    part = MIMEBase('application', "octet-stream")

    #attachment
    part.set_payload(open("Coolist_Flyer.pdf", "rb").read())

    encoders.encode_base64(part)

    part.add_header('Content-Disposition', 'attachment; filename="Coolist_Flyer.pdf"')
    msg.attach(part)

    smtpObj.sendmail(temp_user_name, temp_user_name, msg.as_string())

    print("A copy of this email has been sent to yourself. Please preview it thoroughly...")
    time.sleep(5)
    check = input("If you are ready to send this email to the mailing list, enter 'yes'. Otherwise, enter 'no' to "
                  "exit and edit email")
    while True:
        while check.lower() not in ('yes', 'no'):
            check = input("Invalid response. Please enter 'yes' if everything looks good. Otherwise enter 'no' to exit "
                          "program")
        if check.lower() == 'no':
            print("\nTerminating program..... \nPlease fix the text and restart program.")
            time.sleep(3)
            sys.exit()
        if check.lower() == 'yes':
            break


def load_xlsx(filename):
    if ".xlsx" not in filename:
        filename += '.xlsx'
    workbook = openpyxl.load_workbook(filename)
    wbsheet = workbook.get_sheet_by_name('Sheet1')
    counter = 0
    for row_index in range(1, wbsheet.max_row + 1):
        full_name = str(wbsheet.cell(row_index, 1).value).split()
        first_name = full_name[0]
        if len(full_name) == 2:
            full_name_str = full_name[0] + ' ' + full_name[1]
        if len(full_name) == 3:
            full_name_str = full_name[0] + ' ' + full_name[1] + ' ' + full_name[2]
        add_contact(full_name_str, wbsheet.cell(row_index, 2).value,
                    wbsheet.cell(row_index, 3).value)
        counter += 1
    print(counter, 'contact(s) were recorded')


def print_contact_list():
    fmt = '{:<3} {!s:<20} {!s:<40} {!s:<25}'
    print(fmt.format('', 'Name', 'Email Address', 'Company'))
    index = 1
    for contact in contactList:
        print(fmt.format(index, contact.name, contact.email, contact.company))
        index += 1
    print('\nPrinting contacts list complete.')
    ready = input("Enter 'yes' if everything looks good. Otherwise enter 'no' to exit program")
    while True:
        while ready.lower() not in ('yes', 'no', 'y', 'n'):
            ready = input("Invalid response. Please enter 'yes' if everything looks good. Otherwise enter 'no' to exit "
                          "program")
        if ready.lower() == 'no' or ready.lower() == 'n':
            print("\nTerminating program..... \nPlease fix the problem cells in Excel file and restart program.")
            time.sleep(3)
            sys.exit()
        if ready.lower() == 'yes' or ready.lower() == 'y':
            break


def add_contact(name, email_address, company):
    temp_contact = Contact(name, email_address, company)
    contactList.append(temp_contact)


def main():
    unlock_program()

    excel_file_name = input("What is the name of the excel file")
    load_xlsx(excel_file_name)
    print("\nCheck to make sure there are no errors in the contacts list")
    print_contact_list()

    emailUser = input("\n Login Credentials \nEnter email address: ")
    emailPass = input("Enter email password: ")
    login(emailUser, emailPass)

    amount_sent = send_email(emailUser)

    print(amount_sent, "email(s) were successfully sent. \n\n Program will auto-terminate in 3 seconds...")
    time.sleep(3)

    smtpObj.quit()


main()
# add a note about bold and underline manually
# add catch email sent

# PRIORITY add a test dummy email to self

# if company name variables are needed just add {company} into paragraphs as user
